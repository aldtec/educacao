from django.shortcuts import render
from django.http import HttpResponse
#import os

from annoying.functions import get_object_or_None

from tempfile import NamedTemporaryFile
from openpyxl import load_workbook, Workbook
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Fill
from openpyxl.cell import Cell
from openpyxl.drawing.image import Image

from .models import Docente, Calendario
from django.core import serializers

import datetime
import pendulum
pendulum.set_locale('pt-br')
#os.path.join(BASE_DIR, "static")
# logo = Image("static/image/logo_pequeno.png")
# logo.height = 70
# logo.width = 70
#import calendar

from django.conf.urls.static import static
# Create your views here.

def index(request):
	#excel_file = request.FILES["Ponto_docente.xlsx"]
	excel_filee = "PontoDocente.xlsx"
	wb = openpyxl.load_workbook(excel_filee)
	ws1 = wb.create_sheet("Mysheet")
	wb.save('balances.xlsx')
	
	#latest_question_list = Question.objects.order_by('-pub_date')[:5]
	response = wb.sheetnames #type()
	return HttpResponse(response)

def comeee(request):
	po = "excel/Ponto_docente.xlsx"
	wb = load_workbook(filename=po)
	ponto = "modelo"
	p = wb[ponto]
	#p1 = wb["CAPA1"]
	#prof = Docente.objects.get(pk=1)
	#p['A6'].value = prof.nome
	#p['A1'].value = prof.horario

	wb.save("excel/ponto.xlsx")

	#wb = wb1.copy_worksheet(p)
	return HttpResponse(p['A1'].value + p['A6'].value)

def come(request):
	mes = 4
	ano =2020
	ext = dsc(1, mes, ano)
	po = "excel/Ponto_docente.xlsx"
	wb = load_workbook(filename=po)
	ponto = "modelo"
	ps = wb[ponto]

	# Capa do ponto - falta converter para função
	capa  = "capa"
	cap = wb[capa]
	cap['A11'].value = ext['mes'].upper()
	#cap.add_image(logo, "B2")

	# Abertura do ponto - falta converter para função
	abert = "abertura"
	abt = wb[abert]
	abt['A13'].value = dtlocalhj()['s'] # Local 1º dia do mes

	# Para obter quantidades de dias no mes
	fim_mes = ext['fm']

	#Onde a linha começa no excel
	numero = 11
	
	#Para setar a cor de fundo das celulas
	fundo_cinza = PatternFill(fill_type='solid', start_color='BFBFBF', end_color='BFBFBF')#PatternFill(fill_type=None, start_color='A7A7A7', end_color='A7A7A7')

	ps['S4'].value = str(ext['mes'].capitalize())+" "+str(ext['ano'])

	#treta = dici[0][1].format('ddd')

	i = 1
	while i <= fim_mes:
		linha = 10 + i
		ps['A'+str(linha)].value = i
		week = dsc(i, mes, ano)
		diasfolga = dayoff(mes, ano)
		ps['B'+str(linha)].value = week['sss']
		total_colunas = 20
		init_colunas = 3
		connn = str(pendulum.date(ano, mes, i)) #str(ano)+"-"+str(mes)+"-"+str(i) #ano+"-"+mes+"-"+i

		#pendulum.parse(dt.data.strftime("%Y-%m-%d"))
		elem =  get_object_or_None(Calendario, data=pendulum.date(ano, mes, i))
		if elem:
			#mude o valor da coluna caso necessario
			ps.cell(row=linha, column=19, value=elem.get_observ_display().upper())
			while init_colunas <= total_colunas: 	
				ps.cell(row=linha, column=init_colunas).fill = fundo_cinza
				init_colunas += 1

		init_colunas = 3
		if week['f']: #is True:
			while init_colunas <= total_colunas: 	
				ps.cell(row=linha, column=init_colunas).fill = fundo_cinza
				init_colunas += 1
		i = int(i)
		i += 1


	testando = Docente.objects.all().order_by("rf_vinc")
	dicionario = []
	for cada in testando:
		teacher = Docente.objects.get(nome=cada)
		dicionario.append([
			teacher.nome.upper(), 
			teacher.rf_vinc, 
			teacher.qpe, 
			teacher.cargo,
			teacher.regencia,
			teacher.hor_col,
			teacher.turma,
			teacher.horario,
			teacher.get_jornada_display()
			])

	for cdprof in dicionario:
		tgnome = cdprof[1].replace('/', '-')
		target = wb.copy_worksheet(ps)
		logo = Image("static/image/logo_pequeno.png")
		logo.height = 70
		logo.width = 70
		target.add_image(logo, "B1")
		target.title = tgnome
		tg = wb[tgnome]
		tg['C6'].value = cdprof[0]
		tg['H6'].value = cdprof[1]
		tg['H7'].value = cdprof[2]
		tg['C7'].value = cdprof[3]
		tg['C8'].value = cdprof[4]
		tg['I8'].value = cdprof[5]
		tg['T8'].value = cdprof[6]
		tg['C9'].value = cdprof[7]
		tg['S7'].value = cdprof[8]


	final = "final"
	fl = wb[final]
	fl['A11'].value = dtlocalpx(mes, ano)
	target = wb.copy_worksheet(fl)
	target.title = 'Encerramento'

	wb.remove(wb.get_sheet_by_name(ponto))
	wb.remove(wb.get_sheet_by_name(final))

	wb.save("excel/ponto.xlsx")
	#z = c.soma(2,3)

	return HttpResponse(dayoff(4,2020))#dt.format('dddd Do [de] MMMM [de] YYYY').capitalize())dayoff(4,2020) 

def dsc(dia, mes, ano):
  date = pendulum.date(ano, mes, dia)
  mes = date.format('MMMM')
  ano = date.year
  fundo = ['dom','sáb']
  fmes = date.days_in_month
  x = date.format('ddd')
  if x in fundo:
    f = True
  else:
    f = False
  l = {
    "s"   : x[0].capitalize(),
    "sss" : x.capitalize(),
    "f"   : f,
    "fm"  : fmes,
    "mes" : mes,
    "ano" : ano
  }
  return l 

def dtlocalpx(mes, ano):
  data = pendulum.date(ano, mes, 1)
  fim = data.days_in_month
  x = pendulum.date(ano, mes, fim).add(days=1)
  ano = x.format('YYYY')
  mes = x.format('MMMM').capitalize()
  m = x.format('ddd')
  dia = x.day
  if m == 'dom':
    dia += 1
  elif m == 'sáb':
    dia += 2
  r = "São Paulo, "+str(dia)+" de "+mes+" de "+ano+"."
  return r


def dtlocalhj():
  x = pendulum.now()
  #x = x.add(days=d)
  ano = x.format('YYYY')
  mes = x.format('MMMM').capitalize()
  dia = x.format('Do')
  l = {
  	"s"   : "São Paulo, 1 de "+mes+" de "+ano+".",
  	"pri" : "",
    "r"   : "São Paulo, "+dia+" de "+mes+" de "+ano+"."
  }
  return l

def dayoff(mes, ano):
	#pesqui = pendulum.parse(dtc)#parse(dtc.strftime("%Y-%m-%d"))
	#mes = pesqui.month
	#ano = pesqui.year

	# Filtra com mais de um argumento e ordena o resultado
	filtra = Calendario.objects.filter(data__year = ano).filter(data__month = mes).order_by('data')

	dici = []
	for dt in filtra:
		dici.append([
			dt.descricao.upper(),
			#dt.data.strftime("%a, %d, %m, %Y"),
			#pendulum.parse(dt.data.strftime("%Y-%m-%d")),
			dt.data.strftime("%Y-%m-%d"),
			#dt.data,
			dt.get_observ_display().upper()
			])

	return dici

#folga.data.strftime("%d, %m, %Y"), # REsulta em objeto datetime

	#object_list = serializers.serialize("python", Docente.objects.all())
	# listagem = {} #[] 	docentes = Docente.get.all()
	# for object in object_list:
	# 	listagem.update({object}) 
	# 	#for obj in object['fields'].items():
	# 		#listagem.append(obj)
	#for prof in docentes:  	resultDict = model_to_dict(docentes)
	#	dicio.append(prof)

# for obj in serializers.deserialize('xml', data, handle_forward_references=True):
#     if obj.deferred_fields is not None:
#         objs_with_deferred_fields.append(obj)




# cell_a = 'A'
# i = 1
# while i <= a:
#     i = str(i)
#     cell = cell_a + i
#     work_sheet[cell] = 'Test text'
#     work_sheet_a1 = work_sheet[cell]
#     work_sheet_a1.font = Font(size=i)
#     i = int(i)
#     i = i + 1


# latest_question_list = Question.objects.order_by('-pub_date')[:5]
#    output = ', '.join([q.question_text for q in latest_question_list])workbook.sheetnames
# ['Sheet 1']

# >>> sheet = workbook.active
# >>> sheet
# <Worksheet "Sheet 1">

# >>> sheet.title
# 'Sheet 1'